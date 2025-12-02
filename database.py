# database.py
import os
import sqlite3
import json
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import config # وارد کردن کل ماژول config

class DatabaseManager:
    def __init__(self):
        self.conn = sqlite3.connect(config.DATABASE_NAME)
        self.cursor = self.conn.cursor()
        self._create_tables()

    def _create_tables(self):
        # جدول اصلی رزومه‌ها: ستون is_blocked برای قابلیت بلاک/آنبلاک اضافه شد
        self.cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS resumes (
                user_id INTEGER PRIMARY KEY,
                {', '.join([f'{field} TEXT' for field in config.RESUME_FIELDS])},
                is_admin_notified INTEGER DEFAULT 0,
                is_blocked INTEGER DEFAULT 0,
                is_deleted INTEGER DEFAULT 0,
                deleted_at TEXT,
                deleted_by INTEGER
            )
        """)
        # جدول لاگ فعالیت‌ها
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                level TEXT,
                message TEXT
            )
        """)
        self.conn.commit()
        # جدول ثبت اعمال ادمین (تاریخچه تغییرات)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS admin_actions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                admin_id INTEGER,
                target_user_id INTEGER,
                action_type TEXT,
                field_name TEXT,
                old_value TEXT,
                new_value TEXT
            )
        """)
        self.conn.commit()
        # Ensure all fields from config.RESUME_FIELDS exist as columns (migrate if needed)
        try:
            self.cursor.execute("PRAGMA table_info(resumes)")
            existing_cols = [row[1] for row in self.cursor.fetchall()]
            for field in config.RESUME_FIELDS:
                if field not in existing_cols:
                    self.cursor.execute(f"ALTER TABLE resumes ADD COLUMN {field} TEXT")
            # ensure new admin/soft-delete columns exist
            extra_cols = ['is_admin_notified', 'is_blocked', 'is_deleted', 'deleted_at', 'deleted_by']
            for c in extra_cols:
                if c not in existing_cols:
                    try:
                        self.cursor.execute(f"ALTER TABLE resumes ADD COLUMN {c} TEXT")
                    except Exception:
                        pass
            self.conn.commit()
        except Exception:
            # If pragma/alter not supported or fails, ignore and continue (table already created earlier)
            pass

    def log(self, level, message):
        """ثبت رویداد در دیتابیس و فایل متنی"""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.cursor.execute(
            "INSERT INTO logs (timestamp, level, message) VALUES (?, ?, ?)",
            (timestamp, level, message)
        )
        self.conn.commit()
        # ثبت در فایل متنی
        with open(config.LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(f"[{timestamp}] [{level}] {message}\n")

    def save_resume_data(self, user_id, data: dict):
        """ذخیره یا به‌روزرسانی اطلاعات رزومه کاربر"""
        # For any list or dict values (e.g., skills, uploaded_files), store as JSON text
        fields = config.RESUME_FIELDS.copy()
        values = []
        for k in fields:
            v = data.get(k)
            if isinstance(v, (list, dict)):
                try:
                    values.append(json.dumps(v, ensure_ascii=False))
                except Exception:
                    values.append(str(v))
            else:
                values.append(v)

        field_placeholders = ', '.join(fields)
        value_placeholders = ', '.join(['?' for _ in fields])

        query = f"""
            INSERT OR REPLACE INTO resumes (user_id, {field_placeholders})
            VALUES (?, {value_placeholders})
        """

        # Prepare params: user_id first
        params = (user_id, *values)
        self.cursor.execute(query, params)
        self.conn.commit()
        self.log("INFO", f"Resume data updated for User ID: {user_id}")
        
    def get_resume_data(self, user_id):
        """دریافت تمام اطلاعات یک کاربر"""
        self.cursor.execute("SELECT * FROM resumes WHERE user_id = ?", (user_id,))
        row = self.cursor.fetchone()
        if row:
            columns = [col[0] for col in self.cursor.description]
            data = dict(zip(columns, row))
            if 'skills' in data and data['skills']:
                try:
                    data['skills'] = json.loads(data['skills'])
                except json.JSONDecodeError:
                    data['skills'] = []
            return data
        return None

    def get_resumes_for_export(self):
        self.cursor.execute("SELECT * FROM resumes")
        return self.cursor.fetchall(), [col[0] for col in self.cursor.description]

    def search_resumes(self, term: str, limit: int = 10, offset: int = 0, filters: dict = None):
        """Search resumes by full_name, username or major with pagination. Returns (rows, total_count)."""
        filters = filters or {}
        like = f"%{term.strip()}%"
        # base where clause - exclude deleted by default
        include_deleted = filters.pop('_include_deleted', False)
        if include_deleted:
            where = "WHERE 1=1"
        else:
            where = "WHERE is_deleted = 0"
        params = []
        if term:
            where += " AND (full_name LIKE ? OR username LIKE ? OR major LIKE ?)"
            params.extend([like, like, like])

        # basic filters support
        if 'study_status' in filters:
            where += " AND study_status = ?"
            params.append(filters['study_status'])
        if 'degree' in filters:
            where += " AND degree = ?"
            params.append(filters['degree'])

        # total count
        count_q = f"SELECT COUNT(user_id) FROM resumes {where}"
        self.cursor.execute(count_q, tuple(params))
        total = self.cursor.fetchone()[0]

        q = f"SELECT user_id, full_name, username, register_date FROM resumes {where} ORDER BY register_date DESC LIMIT ? OFFSET ?"
        exec_params = tuple(params + [limit, offset])
        self.cursor.execute(q, exec_params)
        rows = self.cursor.fetchall()
        return rows, total

    def log_admin_action(self, admin_id: int, target_user_id: int, action_type: str, field_name: str = None, old_value: str = None, new_value: str = None):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.cursor.execute(
            "INSERT INTO admin_actions (timestamp, admin_id, target_user_id, action_type, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (timestamp, admin_id, target_user_id, action_type, field_name, old_value, new_value)
        )
        self.conn.commit()
        self.log("ADMIN", f"Admin {admin_id} {action_type} on user {target_user_id} field {field_name} -> {new_value}")

    def soft_delete_user(self, user_id: int, admin_id: int) -> bool:
        """Mark a user as deleted (soft delete)."""
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.cursor.execute("UPDATE resumes SET is_deleted = 1, deleted_at = ?, deleted_by = ? WHERE user_id = ?", (ts, admin_id, user_id))
            self.conn.commit()
            self.log_admin_action(admin_id, user_id, 'soft_delete', None, None, None)
            return True
        except Exception as e:
            self.log("ERROR", f"soft_delete_user failed: {e}")
            return False

    def restore_user(self, user_id: int, admin_id: int) -> bool:
        """Restore a soft-deleted user."""
        try:
            self.cursor.execute("UPDATE resumes SET is_deleted = 0, deleted_at = NULL, deleted_by = NULL WHERE user_id = ?", (user_id,))
            self.conn.commit()
            self.log_admin_action(admin_id, user_id, 'restore', None, None, None)
            return True
        except Exception as e:
            self.log("ERROR", f"restore_user failed: {e}")
            return False
    
# database.py (فقط تابع export_to_excel اصلاح شده)

    def export_to_excel(self):
        """(مورد 3) تهیه خروجی اکسل از تمام رزومه‌ها"""
        rows, columns = self.get_resumes_for_export()
        if not rows:
            return False, "دیتابیس خالی است."
        
        df = pd.DataFrame(rows, columns=columns)
        
        if 'skills' in df.columns:
             # یک تابع کمکی برای تبدیل JSON امن تعریف می‌کنیم
             def format_skills(x):
                 if x and isinstance(x, str) and x.strip().startswith('['):
                     try:
                         skills_list = json.loads(x)
                         return "\n".join([f"{item['name']}: {item['level']}" for item in skills_list])
                     except json.JSONDecodeError:
                         return "خطای تبدیل JSON"
                 return x # برگرداندن مقدار اصلی در صورت خالی بودن، نبودن رشته یا خطا
                 
             df['skills'] = df['skills'].apply(format_skills)
        # format uploaded_files column (JSON list of paths) to friendly list of filenames
        if 'uploaded_files' in df.columns:
            def format_uploaded(x):
                if x and isinstance(x, str) and x.strip().startswith('['):
                    try:
                        files = json.loads(x)
                        if isinstance(files, list):
                            return "\n".join([os.path.basename(f) for f in files])
                    except json.JSONDecodeError:
                        return "خطای تبدیل JSON"
                return x
            df['uploaded_files'] = df['uploaded_files'].apply(format_uploaded)
        
        # Reorder columns to a predictable export order: user_id first, then RESUME_FIELDS,
        # then any admin/internal flags.
        ordered_fields = ['user_id'] + config.RESUME_FIELDS + ['is_admin_notified', 'is_blocked']
        ordered_existing = [c for c in ordered_fields if c in df.columns]

        if ordered_existing:
            df = df[ordered_existing]

        # Map column headers to Persian labels for a friendly export using RESUME_FIELDS_PERSIAN
        # Build mapping for the fields we expect
        header_map = {}
        # user-friendly labels for extra/internal columns
        extra_labels = {
            'user_id': 'آیدی کاربر',
            'is_admin_notified': 'اطلاع‌رسانی ادمین',
            'is_blocked': 'بلاک'
        }

        for col in df.columns:
            if col in extra_labels:
                header_map[col] = extra_labels[col]
            else:
                # find index in RESUME_FIELDS to pick corresponding Persian label
                if col in config.RESUME_FIELDS:
                    idx = config.RESUME_FIELDS.index(col)
                    # fall back to FIELD_LABELS if RESUME_FIELDS_PERSIAN missing
                    try:
                        header_map[col] = config.RESUME_FIELDS_PERSIAN[idx]
                    except Exception:
                        header_map[col] = config.FIELD_LABELS.get(col, col)
                else:
                    header_map[col] = config.FIELD_LABELS.get(col, col)

        df = df.rename(columns=header_map)

        try:
            df.to_excel(config.EXCEL_OUTPUT, index=False, engine='openpyxl')

            # Post-process with openpyxl to improve column widths, wrap text and header style
            try:
                wb = load_workbook(config.EXCEL_OUTPUT)
                ws = wb.active

                # Bold header row and set alignment
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Adjust column widths based on max length in each column (header + cells)
                for i, col in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(i)
                    # Compute max length of values in this column
                    try:
                        series = df[col].astype(str)
                        max_length = max(series.map(len).max(), len(str(col)))
                    except Exception:
                        max_length = len(str(col))

                    # Set a reasonable width (cap to avoid extremely wide columns)
                    adjusted_width = min(max_length * 1.2 + 2, 60)
                    ws.column_dimensions[col_letter].width = adjusted_width

                # Optionally set a default row height for readability
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if not cell.alignment:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                wb.save(config.EXCEL_OUTPUT)
            except Exception as e:
                # Non-fatal: if post-processing fails, still return the generated file
                self.log('ERROR', f'Excel post-processing failed: {e}')

            self.log("INFO", f"Data exported to {config.EXCEL_OUTPUT}")
            return True, config.EXCEL_OUTPUT
        except Exception as e:
            self.log("ERROR", f"Failed to export Excel: {e}")
            return False, f"خطای سیستمی هنگام ساخت فایل اکسل: {e}"

    # ... (بقیه توابع DatabaseManager) ...

    # ===============================================
    #           توابع مورد نیاز پنل ادمین
    # ===============================================

    def get_user_by_search_term(self, term):
        """(مورد 1) جستجو بر اساس نام کامل، بخشی از نام یا یوزرنیم"""
        search_term = f'%{term.strip()}%'
        self.cursor.execute(
            "SELECT user_id, full_name, username FROM resumes WHERE full_name LIKE ? OR username LIKE ?",
            (search_term, search_term)
        )
        return self.cursor.fetchall() # (user_id, full_name, username)

    def get_stats(self, today_date_str):
        """(مورد 4) دریافت آمار کلی کاربران"""
        self.cursor.execute("SELECT COUNT(user_id) FROM resumes")
        total_users = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(user_id) FROM resumes WHERE register_date LIKE ?", (f'{today_date_str}%',))
        today_users = self.cursor.fetchone()[0]
        
        return total_users, today_users

    def delete_user(self, user_id):
        """(مورد 5) حذف کاربر از دیتابیس"""
        self.cursor.execute("DELETE FROM resumes WHERE user_id = ?", (user_id,))
        self.conn.commit()
        self.log("ADMIN", f"User {user_id} deleted from database.")
        
    def update_user_field(self, user_id, field_name, new_value):
        """(مورد 6 و 9) ویرایش یک فیلد خاص یا تغییر وضعیت بلاک/آنبلاک"""
        
        allowed_fields = config.RESUME_FIELDS + ['is_blocked', 'is_admin_notified']
        if field_name not in allowed_fields:
            self.log("ERROR", f"Attempted to update invalid field: {field_name}")
            return False
            
        # اگر فیلد مهارت‌ها بود، آن را به JSON تبدیل کن تا ساختار حفظ شود
        if field_name == 'skills':
            # در اینجا فرض بر این است که ادمین هنگام ویرایش، رشته JSON معتبری وارد می‌کند
            # یا منطق پیچیده‌تری در bot.py برای ویرایش مهارت‌ها در نظر گرفته شده است.
            # برای سادگی، اگر یک لیست بود به JSON تبدیل شود
            if isinstance(new_value, list):
                new_value = json.dumps(new_value, ensure_ascii=False)
        
        query = f"UPDATE resumes SET {field_name} = ? WHERE user_id = ?"
        self.cursor.execute(query, (new_value, user_id))
        self.conn.commit()
        self.log("ADMIN", f"User {user_id} field '{field_name}' updated to '{new_value}'.")
        return True

    def get_all_logs(self):
        """(مورد 10) دریافت آخرین لاگ‌های فعالیت"""
        self.cursor.execute("SELECT * FROM logs ORDER BY timestamp DESC LIMIT 500")
        return self.cursor.fetchall()
        
    def close(self):
        self.conn.close()